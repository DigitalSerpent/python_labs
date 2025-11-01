import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import sys
import os

sys.path.append(os.path.join(os.path.dirname(__file__), '..'))
from lib.io_helpers import read_csv
def csv_to_xlsx(csv_path: str, xlsx_path: str) -> None:
    """
    конвертирует CSV в XLSX
    использовать openpyxl
    колонки не менее 8 символов
    """
    #чтение и валидация CSV
    fieldnames, rows = read_csv(csv_path)
    #создание директории для выходного файла, если её нет
    output_path = Path(xlsx_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    #создание Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "лист1"
    #запись заголовков
    ws.append(fieldnames)
    # Запись данных
    for row in rows:
        ws.append([row.get(field, '') for field in fieldnames])
    #настраиваю ширину колонок
    for col_num, column_title in enumerate(fieldnames, 1):
        max_length = 0
        column_letter = get_column_letter(col_num)
        #длина заголовка
        max_length = max(max_length, len(str(column_title)))
        #длина данных в колонке
        for row_num in range(2, ws.max_row + 1):
            cell_value = ws[f"{column_letter}{row_num}"].value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        #ширина 8 символов
        adjusted_width = max(8, max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    #сохраняю
    wb.save(output_path)